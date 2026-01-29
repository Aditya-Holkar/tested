# modules/report_generator.py - Enhanced report generation functions
import json
import csv
import pandas as pd
from datetime import datetime
from io import StringIO, BytesIO
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ReportGenerator:
    """Generate comprehensive reports in various formats with detailed sheets"""
    
    def generate_report(self, data, format='json'):
        """Generate report in specified format"""
        if format == 'json':
            return self._generate_json_report(data)
        elif format == 'csv':
            return self._generate_csv_summary(data)
        elif format == 'html':
            return self._generate_html_report(data)
        elif format == 'excel':
            return self._generate_excel_report(data)
        elif format == 'detailed_excel':
            return self._generate_detailed_excel_report(data)
        else:
            return self._generate_json_report(data)
    
    def _generate_json_report(self, data):
        """Generate JSON report"""
        return json.dumps(data, indent=2, default=str)
    
    def _generate_csv_summary(self, data):
        """Generate CSV summary report"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Website Comprehensive Test Report'])
        writer.writerow(['Generated:', data.get('timestamp', '')])
        writer.writerow([])
        
        # Write summary
        if 'summary' in data:
            writer.writerow(['Summary Statistics'])
            writer.writerow(['Metric', 'Value'])
            for key, value in data['summary'].items():
                writer.writerow([key, value])
            writer.writerow([])
        
        return output.getvalue()
    
    def _generate_html_report(self, data):
        """Generate HTML report"""
        # (Keep the existing HTML generation code)
        html = """<!DOCTYPE html>
<html>
<head>
    <title>Website Test Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; }
        .header { text-align: center; border-bottom: 2px solid #3498db; padding-bottom: 20px; }
        .section { margin: 30px 0; }
        table { width: 100%; border-collapse: collapse; margin: 15px 0; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #3498db; color: white; }
        .status-pass { background-color: #d4edda; }
        .status-fail { background-color: #f8d7da; }
        .status-warning { background-color: #fff3cd; }
    </style>
</head>
<body>
    <div class="header">
        <h1>Website Comprehensive Test Report</h1>
        <p>Generated: """ + data.get('timestamp', '') + """</p>
    </div>
    
    <div class="section">
        <h2>Summary</h2>
        <table>
            <tr><th>Metric</th><th>Value</th></tr>
"""
        
        if 'summary' in data:
            for key, value in data['summary'].items():
                html += f'<tr><td>{key}</td><td>{value}</td></tr>'
        
        html += """
        </table>
    </div>
</body>
</html>
"""
        
        return html
    
    def _generate_excel_report(self, data):
        """Generate basic Excel report"""
        try:
            # Create a BytesIO object to store Excel file
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Sheet 1: Summary
                if 'summary' in data:
                    summary_df = pd.DataFrame(list(data['summary'].items()), columns=['Metric', 'Value'])
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Sheet 2: All Test Cases
                if 'test_cases' in data and data['test_cases']:
                    test_cases_df = pd.DataFrame(data['test_cases'])
                    test_cases_df.to_excel(writer, sheet_name='All Test Cases', index=False)
            
            return output.getvalue()
            
        except Exception as e:
            return f"Error generating Excel report: {str(e)}"
    
    def _generate_detailed_excel_report(self, data):
        """Generate detailed Excel report with multiple sheets for each test type"""
        try:
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                workbook = writer.book
                
                # ============================================================
                # SHEET 1: EXECUTIVE SUMMARY
                # ============================================================
                if 'summary' in data:
                    summary_data = []
                    summary = data['summary']
                    
                    summary_data.append(['EXECUTIVE SUMMARY', ''])
                    summary_data.append(['Generated', data.get('timestamp', '')])
                    summary_data.append([])
                    summary_data.append(['OVERALL STATISTICS', ''])
                    
                    for key, value in summary.items():
                        display_name = key.replace('_', ' ').title()
                        summary_data.append([display_name, value])
                    
                    summary_data.append([])
                    summary_data.append(['TEST BREAKDOWN', ''])
                    
                    # Count test cases by type
                    if 'test_cases' in data and data['test_cases']:
                        test_types = {}
                        for tc in data['test_cases']:
                            test_type = tc.get('Test Type', 'Unknown')
                            test_types[test_type] = test_types.get(test_type, 0) + 1
                        
                        for test_type, count in test_types.items():
                            summary_data.append([test_type, count])
                    
                    summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
                    summary_df.to_excel(writer, sheet_name='Executive Summary', index=False, header=False)
                
                # ============================================================
                # SHEET 2: ALL TEST CASES (DETAILED)
                # ============================================================
                if 'test_cases' in data and data['test_cases']:
                    # Get all test cases
                    all_test_cases = data['test_cases']
                    
                    # Convert to DataFrame with all columns
                    test_cases_data = []
                    for tc in all_test_cases:
                        row = {}
                        # Include all standard columns
                        standard_cols = [
                            'Test ID', 'Module', 'Test Links/Data', 'Test Case Description',
                            'Pre-Conditions', 'Test Steps', 'Expected Result', 'Actual Result',
                            'Status', 'Severity', 'Case Pass/Fail', 'Comments/Bug ID', 'Resolutions',
                            'Test Type', 'Timestamp'
                        ]
                        
                        for col in standard_cols:
                            row[col] = tc.get(col, '')
                        
                        test_cases_data.append(row)
                    
                    if test_cases_data:
                        test_cases_df = pd.DataFrame(test_cases_data)
                        test_cases_df.to_excel(writer, sheet_name='All Test Cases', index=False)
                
                # ============================================================
                # SHEET 3: LINK STATUS RESULTS
                # ============================================================
                if 'link_results' in data and data['link_results']:
                    link_data = []
                    for result in data['link_results']:
                        row = {
                            'URL': result.get('url', ''),
                            'Status Code': result.get('status_code', ''),
                            'Status Text': result.get('status_text', ''),
                            'Category': result.get('status_category', ''),
                            'Response Time (ms)': result.get('response_time_ms', ''),
                            'Final URL': result.get('final_url', ''),
                            'Timestamp': result.get('timestamp', '')
                        }
                        link_data.append(row)
                    
                    if link_data:
                        link_df = pd.DataFrame(link_data)
                        link_df.to_excel(writer, sheet_name='Link Status Results', index=False)
                
                # ============================================================
                # SHEET 4: PERFORMANCE ANALYSIS
                # ============================================================
                if 'performance_results' in data and data['performance_results']:
                    perf_data = []
                    for result in data['performance_results']:
                        if isinstance(result, dict):
                            row = {
                                'Test ID': result.get('Test ID', ''),
                                'Module': result.get('Module', ''),
                                'URL': result.get('Test Links/Data', ''),
                                'Description': result.get('Test Case Description', ''),
                                'Actual Result': result.get('Actual Result', ''),
                                'Status': result.get('Status', ''),
                                'Severity': result.get('Severity', ''),
                                'Comments': result.get('Comments/Bug ID', ''),
                                'Resolutions': result.get('Resolutions', '')
                            }
                            perf_data.append(row)
                    
                    if perf_data:
                        perf_df = pd.DataFrame(perf_data)
                        perf_df.to_excel(writer, sheet_name='Performance Analysis', index=False)
                
                # ============================================================
                # SHEET 5: ACCESSIBILITY TESTING
                # ============================================================
                if 'accessibility_results' in data and data['accessibility_results']:
                    acc_data = []
                    for result in data['accessibility_results']:
                        if isinstance(result, dict):
                            row = {
                                'Test ID': result.get('Test ID', ''),
                                'Module': result.get('Module', ''),
                                'URL': result.get('Test Links/Data', ''),
                                'Description': result.get('Test Case Description', ''),
                                'Actual Result': result.get('Actual Result', ''),
                                'Status': result.get('Status', ''),
                                'Severity': result.get('Severity', ''),
                                'WCAG Level': self._get_wcag_level(result.get('Module', '')),
                                'Resolutions': result.get('Resolutions', '')
                            }
                            acc_data.append(row)
                    
                    if acc_data:
                        acc_df = pd.DataFrame(acc_data)
                        acc_df.to_excel(writer, sheet_name='Accessibility Testing', index=False)
                
                # ============================================================
                # SHEET 6: SEO ANALYSIS
                # ============================================================
                if 'seo_results' in data and data['seo_results']:
                    seo_data = []
                    for result in data['seo_results']:
                        if isinstance(result, dict):
                            row = {
                                'Test ID': result.get('Test ID', ''),
                                'Module': result.get('Module', ''),
                                'URL': result.get('Test Links/Data', ''),
                                'Description': result.get('Test Case Description', ''),
                                'Actual Result': result.get('Actual Result', ''),
                                'Status': result.get('Status', ''),
                                'Severity': result.get('Severity', ''),
                                'Resolutions': result.get('Resolutions', '')
                            }
                            seo_data.append(row)
                    
                    if seo_data:
                        seo_df = pd.DataFrame(seo_data)
                        seo_df.to_excel(writer, sheet_name='SEO Analysis', index=False)
                
                # ============================================================
                # SHEET 7: BUTTON TESTING (ENHANCED)
                # ============================================================
                if 'button_results' in data and data['button_results']:
                    button_data = []
                    for result in data['button_results']:
                        if isinstance(result, dict):
                            row = {
                                'Test ID': result.get('Test ID', ''),
                                'Module': result.get('Module', ''),
                                'URL': result.get('Test Links/Data', ''),
                                'Description': result.get('Test Case Description', ''),
                                'Test Steps': result.get('Test Steps', ''),
                                'Expected Result': result.get('Expected Result', ''),
                                'Actual Result': result.get('Actual Result', ''),
                                'Status': result.get('Status', ''),
                                'Severity': result.get('Severity', ''),
                                'Case Pass/Fail': result.get('Case Pass/Fail', ''),
                                'Comments': result.get('Comments/Bug ID', ''),
                                'Resolutions': result.get('Resolutions', ''),
                                'Function Names': result.get('function_names', ''),  # New column
                                'Redirected URL': result.get('redirected_url', '')  # New column
                            }
                            button_data.append(row)
                    
                    if button_data:
                        button_df = pd.DataFrame(button_data)
                        button_df.to_excel(writer, sheet_name='Button Testing', index=False)       
                # ============================================================
                # SHEET 8: SPELLING CHECK
                # ============================================================
                if 'spelling_results' in data and data['spelling_results']:
                    spelling_data = []
                    for result in data['spelling_results']:
                        if isinstance(result, dict):
                            row = {
                                'Test ID': result.get('Test ID', ''),
                                'Module': result.get('Module', ''),
                                'URL': result.get('Test Links/Data', ''),
                                'Description': result.get('Test Case Description', ''),
                                'Actual Result': result.get('Actual Result', ''),
                                'Status': result.get('Status', ''),
                                'Severity': result.get('Severity', ''),
                                'Resolutions': result.get('Resolutions', '')
                            }
                            spelling_data.append(row)
                    
                    if spelling_data:
                        spelling_df = pd.DataFrame(spelling_data)
                        spelling_df.to_excel(writer, sheet_name='Spelling Check', index=False)
                
                # ============================================================
                # SHEET 9: FONT ANALYSIS
                # ============================================================
                if 'font_results' in data and data['font_results']:
                    font_data = []
                    for result in data['font_results']:
                        if isinstance(result, dict):
                            row = {
                                'Test ID': result.get('Test ID', ''),
                                'Module': result.get('Module', ''),
                                'URL': result.get('Test Links/Data', ''),
                                'Description': result.get('Test Case Description', ''),
                                'Actual Result': result.get('Actual Result', ''),
                                'Status': result.get('Status', ''),
                                'Severity': result.get('Severity', ''),
                                'Resolutions': result.get('Resolutions', '')
                            }
                            font_data.append(row)
                    
                    if font_data:
                        font_df = pd.DataFrame(font_data)
                        font_df.to_excel(writer, sheet_name='Font Analysis', index=False)
                # ============================================================
                # SHEET 14: CLICK EVENT ANALYSIS (NEW SHEET)
                # ============================================================
                if 'button_click_summary' in data and data['button_click_summary']:
                    click_data = []
                    summary = data['button_click_summary']
                    
                    # Summary section
                    click_data.append(['CLICK EVENT TESTING SUMMARY', ''])
                    click_data.append(['Generated', data.get('timestamp', '')])
                    click_data.append([])
                    click_data.append(['OVERALL STATISTICS', ''])
                    click_data.append(['Total Buttons Tested', summary.get('total_buttons_tested', 0)])
                    click_data.append(['Buttons with Click Events', summary.get('buttons_with_click_events', 0)])
                    click_data.append(['Buttons without Click Events', summary.get('buttons_without_click_events', 0)])
                    click_data.append(['Buttons with Functions', summary.get('buttons_with_functions', 0)])
                    click_data.append(['Buttons with Redirects', summary.get('buttons_with_redirects', 0)])
                    click_data.append([])
                    
                    # Detailed results
                    click_data.append(['DETAILED CLICK TEST RESULTS', ''])
                    click_data.append(['Button ID', 'Button Text', 'Click Will Execute', 
                                      'Function Names', 'Redirected URL', 'Execution Details'])
                    
                    for result in summary.get('click_test_results', []):
                        click_data.append([
                            result.get('button_id', ''),
                            result.get('button_text', '')[:50],
                            'Yes' if result.get('will_execute') else 'No',
                            ', '.join(result.get('function_names', [])),
                            result.get('redirected_url', ''),
                            result.get('execution_details', '')[:100]
                        ])
                    
                    click_data.append([])
                    click_data.append(['RECOMMENDATIONS', ''])
                    click_data.append(['1. Buttons without click events', 
                                      'Consider adding functionality or remove if not needed'])
                    click_data.append(['2. Buttons without proper redirects', 
                                      'Ensure form submissions and links have valid destinations'])
                    click_data.append(['3. Inline JavaScript functions', 
                                      'Consider moving to external JavaScript files for maintainability'])
                    click_data.append(['4. Function naming', 
                                      'Use descriptive function names for better code readability'])
                    
                    # Create DataFrame
                    click_df = pd.DataFrame(click_data, columns=['Category', 'Value/Details'])
                    click_df.to_excel(writer, sheet_name='Click Event Analysis', index=False, header=False)
                
                # ============================================================
                # SHEET 15: FUNCTION ANALYSIS (NEW SHEET)
                # ============================================================
                if 'button_results' in data and data['button_results']:
                    function_data = []
                    function_map = {}
                    
                    # Analyze functions used across all buttons
                    for result in data['button_results']:
                        if isinstance(result, dict):
                            functions = result.get('function_names', '')
                            if functions and functions != 'None':
                                func_list = [f.strip() for f in functions.split(',')]
                                for func in func_list:
                                    function_map[func] = function_map.get(func, 0) + 1
                    
                    if function_map:
                        function_data.append(['FUNCTION USAGE ANALYSIS', ''])
                        function_data.append(['Generated', data.get('timestamp', '')])
                        function_data.append([])
                        function_data.append(['Function Name', 'Usage Count', 'Recommendation'])
                        
                        for func, count in sorted(function_map.items(), key=lambda x: x[1], reverse=True):
                            recommendation = ''
                            if count > 5:
                                recommendation = 'High usage - ensure function is well-tested'
                            elif 'submit' in func.lower() or 'save' in func.lower():
                                recommendation = 'Critical function - validate data handling'
                            elif 'delete' in func.lower() or 'remove' in func.lower():
                                recommendation = 'Destructive action - add confirmation'
                            else:
                                recommendation = 'Standard function'
                            
                            function_data.append([func, count, recommendation])
                        
                        function_df = pd.DataFrame(function_data, columns=['Metric', 'Value', 'Additional'])
                        function_df.to_excel(writer, sheet_name='Function Analysis', index=False, header=False)                
                # ============================================================
                # SHEET 10: RESPONSIVENESS CHECK
                # ============================================================
                if 'responsiveness_results' in data and data['responsiveness_results']:
                    responsive_data = []
                    for result in data['responsiveness_results']:
                        if isinstance(result, dict):
                            row = {
                                'Test ID': result.get('Test ID', ''),
                                'Module': result.get('Module', ''),
                                'URL': result.get('Test Links/Data', ''),
                                'Description': result.get('Test Case Description', ''),
                                'Actual Result': result.get('Actual Result', ''),
                                'Status': result.get('Status', ''),
                                'Severity': result.get('Severity', ''),
                                'Resolutions': result.get('Resolutions', '')
                            }
                            responsive_data.append(row)
                    
                    if responsive_data:
                        responsive_df = pd.DataFrame(responsive_data)
                        responsive_df.to_excel(writer, sheet_name='Responsiveness Check', index=False)
                
                # ============================================================
                # SHEET 11: BROWSER COMPATIBILITY
                # ============================================================
                if 'browser_compatibility_results' in data and data['browser_compatibility_results']:
                    browser_data = []
                    for result in data['browser_compatibility_results']:
                        if isinstance(result, dict):
                            row = {
                                'Test ID': result.get('Test ID', ''),
                                'Module': result.get('Module', ''),
                                'URL': result.get('Test Links/Data', ''),
                                'Description': result.get('Test Case Description', ''),
                                'Actual Result': result.get('Actual Result', ''),
                                'Status': result.get('Status', ''),
                                'Severity': result.get('Severity', ''),
                                'Resolutions': result.get('Resolutions', '')
                            }
                            browser_data.append(row)
                    
                    if browser_data:
                        browser_df = pd.DataFrame(browser_data)
                        browser_df.to_excel(writer, sheet_name='Browser Compatibility', index=False)
                
                # ============================================================
                # SHEET 12: TEST CASE STATISTICS
                # ============================================================
                if 'test_cases' in data and data['test_cases']:
                    # Calculate statistics
                    stats_data = []
                    
                    # Overall statistics
                    total_cases = len(data['test_cases'])
                    passed_cases = sum(1 for tc in data['test_cases'] if tc.get('Case Pass/Fail') == 'Pass')
                    failed_cases = total_cases - passed_cases
                    pass_rate = (passed_cases / total_cases * 100) if total_cases > 0 else 0
                    
                    stats_data.append(['OVERALL STATISTICS', ''])
                    stats_data.append(['Total Test Cases', total_cases])
                    stats_data.append(['Passed Cases', passed_cases])
                    stats_data.append(['Failed Cases', failed_cases])
                    stats_data.append(['Pass Rate', f'{pass_rate:.2f}%'])
                    stats_data.append([])
                    
                    # Statistics by test type
                    stats_data.append(['STATISTICS BY TEST TYPE', ''])
                    test_type_stats = {}
                    for tc in data['test_cases']:
                        test_type = tc.get('Test Type', 'Unknown')
                        if test_type not in test_type_stats:
                            test_type_stats[test_type] = {'total': 0, 'passed': 0}
                        test_type_stats[test_type]['total'] += 1
                        if tc.get('Case Pass/Fail') == 'Pass':
                            test_type_stats[test_type]['passed'] += 1
                    
                    for test_type, stats in test_type_stats.items():
                        type_pass_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
                        stats_data.append([test_type, f"{stats['total']} cases ({type_pass_rate:.1f}% pass rate)"])
                    
                    stats_data.append([])
                    
                    # Statistics by severity
                    stats_data.append(['STATISTICS BY SEVERITY', ''])
                    severity_stats = {}
                    for tc in data['test_cases']:
                        severity = tc.get('Severity', 'Medium')
                        if severity not in severity_stats:
                            severity_stats[severity] = {'total': 0, 'passed': 0}
                        severity_stats[severity]['total'] += 1
                        if tc.get('Case Pass/Fail') == 'Pass':
                            severity_stats[severity]['passed'] += 1
                    
                    for severity in ['Critical', 'High', 'Medium', 'Low', 'Info']:
                        if severity in severity_stats:
                            stats = severity_stats[severity]
                            severity_pass_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
                            stats_data.append([severity, f"{stats['total']} cases ({severity_pass_rate:.1f}% pass rate)"])
                    
                    # Create DataFrame
                    stats_df = pd.DataFrame(stats_data, columns=['Category', 'Value'])
                    stats_df.to_excel(writer, sheet_name='Test Case Statistics', index=False, header=False)
                
                # ============================================================
                # SHEET 13: RECOMMENDATIONS
                # ============================================================
                recommendations_data = [
                    ['PRIORITY RECOMMENDATIONS', ''],
                    ['', ''],
                    ['1. CRITICAL ISSUES', 'Address immediately'],
                    ['• Fix broken links (404 errors)', 'High priority'],
                    ['• Resolve server errors (5xx)', 'Immediate attention required'],
                    ['• Fix accessibility violations', 'Legal compliance'],
                    ['', ''],
                    ['2. HIGH PRIORITY ISSUES', 'Address within 1 week'],
                    ['• Optimize page load performance', 'User experience'],
                    ['• Fix high-severity SEO issues', 'Search ranking'],
                    ['• Resolve browser compatibility', 'Cross-browser support'],
                    ['', ''],
                    ['3. MEDIUM PRIORITY ISSUES', 'Address within 2 weeks'],
                    ['• Improve responsive design', 'Mobile usability'],
                    ['• Fix spelling and grammar', 'Professionalism'],
                    ['• Optimize font usage', 'Readability'],
                    ['', ''],
                    ['4. GENERAL RECOMMENDATIONS', 'Ongoing maintenance'],
                    ['• Regular link checking', 'Monthly'],
                    ['• Performance monitoring', 'Weekly'],
                    ['• SEO optimization', 'Continuous'],
                    ['• Accessibility audits', 'Quarterly'],
                    ['', ''],
                    ['REPORT INFORMATION', ''],
                    ['Report Generated', data.get('timestamp', '')],
                    ['Tool Used', 'Website Comprehensive Tester'],
                    ['Recommendation', 'Review detailed test cases in respective sheets']
                ]
                
                rec_df = pd.DataFrame(recommendations_data, columns=['Recommendation', 'Priority/Action'])
                rec_df.to_excel(writer, sheet_name='Recommendations', index=False, header=False)
                
                # ============================================================
                # Apply formatting to all sheets
                # ============================================================
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if cell.value:
                                    cell_length = len(str(cell.value))
                                    if cell_length > max_length:
                                        max_length = cell_length
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Freeze top row
                    worksheet.freeze_panes = 'A2'
                    
                    # Apply basic formatting to header row
                    if worksheet.max_row > 0 and worksheet.max_column > 0:
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
            
            return output.getvalue()
            
        except Exception as e:
            return f"Error generating detailed Excel report: {str(e)}"
    
    def _create_button_statistics_sheet(self, writer, button_data):
            """Create detailed button statistics sheet"""
            if not button_data:
                return
            
            # Filter out summary rows
            detailed_data = [row for row in button_data if row.get('Test ID') != 'SUMMARY']
            
            if not detailed_data:
                return
            
            stats_data = []
            
            # Overall statistics
            total_buttons = len(detailed_data)
            
            # Status breakdown
            status_counts = {}
            severity_counts = {}
            module_counts = {}
            
            for row in detailed_data:
                status = row.get('Status', 'Unknown')
                severity = row.get('Severity', 'Unknown')
                module = row.get('Module', 'Unknown')
                
                status_counts[status] = status_counts.get(status, 0) + 1
                severity_counts[severity] = severity_counts.get(severity, 0) + 1
                module_counts[module] = module_counts.get(module, 0) + 1
            
            # Prepare statistics
            stats_data.append(['BUTTON TESTING STATISTICS', ''])
            stats_data.append(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            stats_data.append([])
            
            stats_data.append(['OVERALL STATISTICS', ''])
            stats_data.append(['Total Buttons Tested', total_buttons])
            stats_data.append([])
            
            stats_data.append(['STATUS BREAKDOWN', ''])
            for status, count in sorted(status_counts.items()):
                percentage = (count / total_buttons * 100) if total_buttons > 0 else 0
                stats_data.append([status, f'{count} ({percentage:.1f}%)'])
            stats_data.append([])
            
            stats_data.append(['SEVERITY BREAKDOWN', ''])
            for severity in ['Critical', 'High', 'Medium', 'Low', 'Info', 'Warning']:
                if severity in severity_counts:
                    count = severity_counts[severity]
                    percentage = (count / total_buttons * 100) if total_buttons > 0 else 0
                    stats_data.append([severity, f'{count} ({percentage:.1f}%)'])
            stats_data.append([])
            
            stats_data.append(['MODULE BREAKDOWN', ''])
            for module, count in sorted(module_counts.items()):
                percentage = (count / total_buttons * 100) if total_buttons > 0 else 0
                stats_data.append([module, f'{count} ({percentage:.1f}%)'])
            stats_data.append([])
            
            # Common Issues
            stats_data.append(['COMMON ACCESSIBILITY ISSUES', ''])
            accessibility_issues = [row for row in detailed_data 
                                if 'Accessibility' in str(row.get('Module', '')) 
                                and row.get('Status') == 'Fail']
            
            if accessibility_issues:
                issue_types = {}
                for issue in accessibility_issues:
                    desc = issue.get('Description', '')
                    if 'accessible name' in desc.lower():
                        issue_types['Missing accessible name'] = issue_types.get('Missing accessible name', 0) + 1
                    elif 'keyboard' in desc.lower():
                        issue_types['Keyboard accessibility'] = issue_types.get('Keyboard accessibility', 0) + 1
                    elif 'focus' in desc.lower():
                        issue_types['Focus indicator'] = issue_types.get('Focus indicator', 0) + 1
                    elif 'aria' in desc.lower():
                        issue_types['ARIA attributes'] = issue_types.get('ARIA attributes', 0) + 1
                
                for issue_type, count in issue_types.items():
                    stats_data.append([issue_type, count])
            else:
                stats_data.append(['No accessibility issues found', ''])
            
            stats_data.append([])
            
            # Recommendations
            stats_data.append(['RECOMMENDATIONS', ''])
            
            if accessibility_issues:
                stats_data.append(['1. Fix accessibility issues', 'High priority for compliance'])
                stats_data.append(['• Add aria-labels to buttons without text', ''])
                stats_data.append(['• Ensure keyboard navigation works', ''])
                stats_data.append(['• Add visible focus indicators', ''])
            
            failed_buttons = [row for row in detailed_data if row.get('Status') == 'Fail']
            if failed_buttons:
                stats_data.append(['2. Fix functional issues', f'{len(failed_buttons)} buttons need attention'])
            
            # Create DataFrame
            stats_df = pd.DataFrame(stats_data, columns=['Category', 'Value'])
            stats_df.to_excel(writer, sheet_name='Button Statistics', index=False, header=False) 
    
    def _get_wcag_level(self, module):
        """Get WCAG level for accessibility module"""
        module_lower = str(module).lower()
        
        if any(term in module_lower for term in ['color', 'contrast', 'audio', 'caption']):
            return 'WCAG 2.1 AA'
        elif any(term in module_lower for term in ['keyboard', 'focus', 'label', 'alt']):
            return 'WCAG 2.1 A'
        elif any(term in module_lower for term in ['aria', 'semantic', 'language']):
            return 'WCAG 2.1 AA'
        else:
            return 'WCAG 2.1 A'
    
    def generate_executive_summary(self, data):
        """Generate executive summary report"""
        summary = {
            'report_title': 'Website Test - Executive Summary',
            'generation_date': data.get('timestamp', ''),
            'overview': self._generate_overview_text(data),
            'key_findings': self._extract_key_findings(data),
            'recommendations': self._generate_recommendations(data),
            'next_steps': self._generate_next_steps()
        }
        
        return summary
    
    def _generate_overview_text(self, data):
        """Generate overview text for executive summary"""
        if 'summary' not in data:
            return "No summary data available."
        
        summary = data['summary']
        text = f"""
        This comprehensive website test report provides detailed analysis across multiple dimensions 
        including link status, performance, accessibility, SEO, and usability. The testing covered 
        {summary.get('total_urls', 0)} URLs resulting in {summary.get('total_test_cases', 0)} 
        individual test cases with an overall pass rate of {summary.get('pass_rate', 0):.1f}%.
        """
        
        return text.strip()
    
    def _extract_key_findings(self, data):
        """Extract key findings from test results"""
        findings = []
        
        if 'test_cases' in data:
            test_cases = data['test_cases']
            
            # Find critical failures
            critical_failures = [tc for tc in test_cases 
                               if tc.get('Severity') == 'Critical' and tc.get('Status') == 'Fail']
            
            if critical_failures:
                findings.append(f"Found {len(critical_failures)} critical issues that require immediate attention.")
            
            # Find high severity failures
            high_failures = [tc for tc in test_cases 
                           if tc.get('Severity') == 'High' and tc.get('Status') == 'Fail']
            
            if high_failures:
                findings.append(f"Found {len(high_failures)} high-severity issues that should be prioritized.")
            
            # Performance findings
            performance_failures = [tc for tc in test_cases 
                                  if 'Performance' in tc.get('Test Type', '') and tc.get('Status') == 'Fail']
            
            if performance_failures:
                findings.append(f"Found {len(performance_failures)} performance issues affecting user experience.")
            
            # Accessibility findings
            accessibility_failures = [tc for tc in test_cases 
                                    if 'Accessibility' in tc.get('Test Type', '') and tc.get('Status') == 'Fail']
            
            if accessibility_failures:
                findings.append(f"Found {len(accessibility_failures)} accessibility compliance issues.")
        
        if 'summary' in data:
            summary = data['summary']
            if summary.get('pass_rate', 100) < 80:
                findings.append(f"Overall pass rate is {summary.get('pass_rate', 0):.1f}%, indicating significant quality issues.")
            elif summary.get('pass_rate', 100) >= 95:
                findings.append(f"Excellent overall pass rate of {summary.get('pass_rate', 0):.1f}% indicates good website quality.")
        
        if not findings:
            findings.append("No major issues found. The website shows good overall quality across all test categories.")
        
        return findings
    
    def _generate_recommendations(self, data):
        """Generate recommendations based on findings"""
        recommendations = []
        
        if 'test_cases' in data:
            test_cases = data['test_cases']
            
            # Check for common issues and provide recommendations
            link_issues = [tc for tc in test_cases 
                          if 'Link' in tc.get('Test Type', '') and tc.get('Status') == 'Fail']
            
            if link_issues:
                recommendations.append("Fix broken links to improve user experience and SEO rankings.")
            
            performance_issues = [tc for tc in test_cases 
                                if 'Performance' in tc.get('Test Type', '') and tc.get('Status') == 'Fail']
            
            if performance_issues:
                recommendations.append("Optimize page load performance for better user engagement and conversion rates.")
            
            accessibility_issues = [tc for tc in test_cases 
                                  if 'Accessibility' in tc.get('Test Type', '') and tc.get('Status') == 'Fail']
            
            if accessibility_issues:
                recommendations.append("Address accessibility issues to ensure compliance with WCAG guidelines and improve inclusivity.")
            
            seo_issues = [tc for tc in test_cases 
                         if 'SEO' in tc.get('Test Type', '') and tc.get('Status') == 'Fail']
            
            if seo_issues:
                recommendations.append("Implement SEO best practices to improve search engine visibility and organic traffic.")
        
        if not recommendations:
            recommendations.append("Continue regular testing and monitoring to maintain website quality.")
            recommendations.append("Consider implementing automated testing in your development pipeline.")
        
        return recommendations
    
    def _generate_next_steps(self):
        """Generate next steps"""
        return [
            "Review and prioritize issues based on severity and impact",
            "Create action plan with timelines for addressing critical and high-severity issues",
            "Schedule follow-up testing after fixes are implemented",
            "Implement continuous monitoring for key performance indicators",
            "Document lessons learned and update development standards",
            "Consider user testing for subjective aspects not covered by automated tests"
        ]